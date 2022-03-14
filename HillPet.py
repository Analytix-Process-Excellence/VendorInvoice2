import asyncio, aiohttp, os, time
import datetime
from datetime import date
import json
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup as bs
from concurrent.futures import ThreadPoolExecutor
import queue, requests

LIMIT = 3
TIMEOUT = 600  # seconds
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36 Edg/98.0.1108.62'

class HillPet:

    def __init__(self,gui_queue,startdate,enddate):
        self.session = requests.session()
        self.gui_queue = gui_queue
        self.client = self.username = self.password = None
        self.startdate = startdate
        self.enddate = enddate
        self.invoices = []
        self.header = ['Client', 'InvoiceNumber', 'InvoiceDate', 'OrderNumber', 'Description']
        self.xldata = []

    async def login(self):
        url = 'https://account.hillsretailorder.com/accounts.login'
        headers = {
            'authority': 'account.hillsretailorder.com',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'accept': '*/*',
            'sec-ch-ua-mobile': '?0',
            'user-agent': USER_AGENT,
            'sec-ch-ua-platform': '"Windows"',
            'content-type': 'application/x-www-form-urlencoded',
            'origin': 'https://hillsretailorder.com',
            'sec-fetch-site': 'same-site',
            'sec-fetch-mode': 'cors',
            'sec-fetch-dest': 'empty',
            'referer': 'https://hillsretailorder.com/',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        data = {
            'loginID': self.username,
            'password': self.password,
            'sessionExpiration': '0',
            'targetEnv': 'jssdk',
            'include': 'profile,data,emails,subscriptions,preferences,id_token,',
            'includeUserInfo': 'true',
            'loginMode': 'standard',
            'lang': 'en',
            'riskContext': '{"b0":30351,"b1":[57,40,114,80],"b2":6,"b3":[],"b4":3,"b5":1,"b6":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36","b7":[{"name":"PDF Viewer","filename":"internal-pdf-viewer","length":2},{"name":"Chrome PDF Viewer","filename":"internal-pdf-viewer","length":2},{"name":"Chromium PDF Viewer","filename":"internal-pdf-viewer","length":2},{"name":"Microsoft Edge PDF Viewer","filename":"internal-pdf-viewer","length":2},{"name":"WebKit built-in PDF","filename":"internal-pdf-viewer","length":2}],"b8":"11:53:27","b9":-330,"b10":{"state":"prompt"},"b11":false,"b12":{"charging":true,"chargingTime":0,"dischargingTime":null,"level":1},"b13":[null,"1366|768|24",false,true]}',
            'APIKey': '3_W8nN0FXgsnmPFRaZ-tN0HqaekeBtySuaaMGLEWyUQ2GNHb8oW61CMgTnk6rBRzrr',
            'source': 'showScreenSet',
            'sdk': 'js_latest',
            'authMode': 'cookie',
            'pageURL': 'https://hillsretailorder.com/login',
            'sdkBuild': '12833',
            'format': 'json'
        }
        async with self.sema:
            async with self.session.post(url,headers=headers,data=data) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                soup = dict(json.loads(content))
                status = soup['statusCode']
                self.UID = soup['UID']
                self.UIDS = soup['UIDSignature']
                self.ID = soup['id_token']
                if status == 200:
                    return True
                else:
                    return False

    async def get_tokens(self):
        url = 'https://api.hillsretailorder.com/authorizationserver/oauth/token'

        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'Accept': 'application/json, text/plain, */*',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
            'sec-ch-ua-platform': '"Windows"',
            'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8',
            'Origin': 'https://hillsretailorder.com',
            'Sec-Fetch-Site': 'same-site',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://hillsretailorder.com/',
            'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        data = {
            'client_id': 'mobile_android',
            'client_secret': 'secret',
            'grant_type': 'custom',
            'UID': self.UID,
            'UIDSignature': self.UIDS,
            'timeStamp': '1646302491',
            'idToken': self.ID,
            'baseSite': 'hillsUSSite'
        }

        async with self.sema:
            async with self.session.post(url,headers=headers,data=data) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                soup = dict(json.loads(content))
                self.accesstoken = f'{str(soup["token_type"]).capitalize()} {soup["access_token"]}'
                return True


    async def get_invoices(self):
        recfrom = 1
        recto = 8
        url = 'https://api.hillsretailorder.com/rest/v2/hillsUSSite/account/current/orderHistory'
        headers = {
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'sec-ch-ua-mobile': '?0',
            'Authorization': self.accesstoken,
            'X-Anonymous-Consents': '%5B%5D',
            'Accept': 'application/json, text/plain, */*',
            'Referer': 'https://hillsretailorder.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
            'sec-ch-ua-platform': '"Windows"',
        }
        flag = True
        while flag:
            params = (
                ('fields', 'DEFAULT'),
                ('recordFrom', recfrom),
                ('recordTo', recto),
                ('orderDateFrom', self.startdate),
                ('orderDateTo', self.enddate),
                ('lang', 'en'),
                ('curr', 'USD'),
            )

            async with self.sema:
                async with self.session.get(url, headers=headers, params=params) as request:
                    response = await request.content.read()
                    content = response.decode('utf-8')
                    soup = dict(json.loads(content))
                    if soup['orders'] is None:
                        self.gui_queue.put({'status': f'No invoices found for client {self.client}'}) if self.gui_queue else None
                        flag = False
                    else:
                        for data in soup['orders']:
                            if ',' in str(data['invoiceNum']):
                                invoiceno = int(str(data['invoiceNum']).count(','))
                                for x in range(invoiceno):
                                    self.invoices.append(
                                            [self.client, data['invoiceNum'].split(',')[x], data['invoiceDate'], data['code']])
                            else:
                                self.invoices.append([data['invoiceNum'],data['invoiceDate'],data['code']])
                        if len(soup['orders']) < 8:
                            flag = False
            recfrom += 8
            recto += 8
        return True

    async def download_invoices(self):
        headers = {
            'Connection': 'keep-alive',
            'Accept': '*/*',
            'Access-Control-Request-Method': 'GET',
            'Authorization': self.accesstoken,
            'Access-Control-Request-Headers': 'authorization,x-anonymous-consents',
            'Origin': 'https://hillsretailorder.com',
            'User-Agent': USER_AGENT,
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://hillsretailorder.com/',
            'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        params = (
            ('lang', 'en'),
            ('curr', 'USD'),
        )
        file_path = os.path.join(os.getcwd(), "Downloads", self.client)

        for invoice in self.invoices:
            url = f'https://api.hillsretailorder.com/rest/v2/hillsUSSite/account/current/invoice-pdf/{invoice[0]}'
            async with self.sema:
                async with self.session.get(url,headers=headers, params=params) as request:
                    response = await request.content.read()
                    if b'The application has encountered an error' not in response:
                        if not os.path.exists(file_path):
                            os.makedirs(file_path)
                        file_name = os.path.join(file_path, f'{invoice[0]}.pdf')
                        with open(file_name, 'wb') as f:
                            f.write(response)
                        self.gui_queue.put(
                            {'status': f'Success : {invoice[0]}'}) if self.gui_queue else None

                        self.xldata.append([self.client, invoice[0], invoice[1], invoice[2], 'File downloaded successfully'])

        return True



    async def download_process(self):
        timeout = aiohttp.ClientTimeout(total=TIMEOUT)
        conn = aiohttp.TCPConnector(limit=5, limit_per_host=5)
        self.sema = asyncio.Semaphore(LIMIT)
        async with aiohttp.ClientSession(connector=conn,timeout=timeout) as self.session:
            login = await self.login()
            if not login:
                self.gui_queue.put({'status': f'Login Error'})if self.gui_queue else None
            tokens = await self.get_tokens()
            if not tokens:
                self.gui_queue.put({'status': f'Tokens expired'}) if self.gui_queue else None
            invoices = await self.get_invoices()
            if not invoices:
                self.gui_queue.put({'status': f'Unable to fetch invoice'})if self.gui_queue else None
            download = await self.download_invoices()
            if not download:
                self.gui_queue.put({'status': f'Unable to download invoice'}) if self.gui_queue else None
            xlupdate = await self.update_xl()
            if not xlupdate:
                self.gui_queue.put({'status': f'Unable to update excel'}) if self.gui_queue else None


    def start_download(self):
        try:
            loop = asyncio.new_event_loop()
            future = asyncio.ensure_future(self.download_process(),loop=loop)
            loop.run_until_complete(future)
            return future.result()
        except Exception as e:
            print(str(e))

    async def update_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(self.header)
        filename = os.path.join(os.getcwd(), "Downloads", 'HillPet.xlsx')
        for xldata in self.xldata:
            ws.append(xldata)
        wb.save(filename)
        return True


class RunHill:

    def __init__(self):
        self.gui_queue = queue.Queue()
# Difference of days between start and end date must not exceed 61

    def run(self,startdate=None,enddate=None):
        run_start = time.perf_counter()
        setting = 'HillPetSettingSheet.xlsx'
        setting_wb = load_workbook(setting, data_only=True, read_only=True)
        setting_ws = setting_wb['Creds'].values
        setting_data = [list(row) for row in setting_ws if row]
        startdate = datetime.datetime.strptime('01/01/2022','%d/%m/%Y').strftime('%Y%m%d')
        enddate = datetime.datetime.strptime('03/03/2022','%d/%m/%Y').strftime('%Y%m%d')
        # print((enddate-startdate).days)
        hill = HillPet(self.gui_queue,startdate,enddate)
        for row_num,row in enumerate(setting_data):
            if len(row) >= 4:
                hill.client = str(row[0]).strip()
                hill.username = str(row[1]).strip()
                hill.password = str(row[2]).strip()
                hill.start_download()

        run_end = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(run_end - run_start)))
        print(f'Time Taken = {time_taken}')
        self.gui_queue.put({"status": f"Time Taken {time_taken}"})


if __name__ == '__main__':
    run = RunHill()
    run.run()