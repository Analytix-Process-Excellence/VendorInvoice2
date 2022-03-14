import os, json, time, asyncio, aiohttp
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup as bs

LIMIT = 3
TIMEOUT = 600  # seconds
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'


class PennInvoice:
    def __init__(self,username,password,client,gui_queue):
        self.username = username
        self.password = password
        self.client = client
        self.gui_queue = gui_queue
        self.invoice_list = []
        self.header = ['Client','InvoiceNumber', 'OrderNumber','OrderDate', 'Description']
        self.xldata = []
        self.login_url = 'https://www.pennvet.com/customer/portal/catalog/customerlogin/!ut/p/z1/04_Sj9CPykssy0xPLMnMz0vMAfIjo8zifS19jTwMvQ38DIyMTA0c3dw9TV39jQwN3A30w9EUGHu4GTj6mVqaBQcaGBkYmOhHEaPfAAdwNCBOPx4FUfiND9ePQrMC0weEzCjIDQ2NMMh0BAD-epEq/p0/IZ7_M9M2H1K0N02250AFGI5EO21082=CZ6_M9M2H1K0N02250AFGI5EO210G0=LA0=Ejavax.portlet.action!processLogin==/'


    async def login(self):
        headers = {
            'authority': 'www.pennvet.com',
            'cache-control': 'max-age=0',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'origin': 'https://www.pennvet.com',
            'upgrade-insecure-requests': '1',
            'dnt': '1',
            'content-type': 'application/x-www-form-urlencoded',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-US,en;q=0.9'
        }

        data = {
            'userId': self.username,
            'password': self.password,
            'submit': 'Login'
        }
        async with self.sema:

            async with self.session.post(self.login_url, headers=headers, data=data) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                login_content = bs(content, "html.parser")
                if login_content:
                    return True




    async def search_invoices(self,startdate,enddate):

        url = 'https://www.pennvet.com/PA_BillingCenter/getInvoicesByDate'
        headers = {
            'authority': 'www.pennvet.com',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'accept': 'application/json, text/javascript, */*; q=0.01',
            'x-requested-with': 'XMLHttpRequest',
            'sec-ch-ua-mobile': '?0',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'cors',
            'sec-fetch-dest': 'empty',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8'
        }

        params = (
            ('customerNumber', '0052650'),
            ('startOrderDate', startdate),
            ('endOrderDate', enddate),
            ('email', self.username),
        )

        async with self.sema:
            # await asyncio.sleep(1)
            async with self.session.get(url, headers=headers, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                invoices = bs(content, "html.parser")
                invoices = json.loads(str(invoices))
                for invoice in invoices:
                    self.invoice_list.append([invoice['invoice'],invoice['orderNumber'],invoice['orderDate']])
        return True

    async def download_invoice(self):
        invoice_url = 'https://www.pennvet.com/customer/myportal/catalog/billing-center/!ut/p/z1/jY_dCoJAFISfxSc4x83VvBQpLVMT__cmFlnMqFVKrMfPrkIoa-4OzHwzBxgUwCQfmpr3TSv5ebxLph980yeu6mEQaiuCkZdRdJyEYKhCPjXgwl2jFVBTjyMkiBqwf_L4RRb-l58xsHl8Dmxa8eGDX4xy3GC8CUs1GDfY9talWmLsTAPiF0OKR7_ntYAyFvxaHTdyaJtK3LJG3KG7pGmBJ9rVivIE2mcIYQ!!/p0/IZ7_M9M2H1K0N81N20QCCJH54T7L97=CZ6_M9M2H1K0NO4E20QKV50GGT20O1=NJviewInvoice=/'

        headers = {
            'authority': 'www.pennvet.com',
            'cache-control': 'max-age=0',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'dnt': '1',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-US,en;q=0.9'
                  }
        for invoiceno in self.invoice_list:
            if invoiceno != 0:
                params = (
                    ('resource', 'viewInvoice'),
                    ('invoiceNum', invoiceno[0]),
                )
                async with self.sema:
                    # await asyncio.sleep(1)
                    async with self.session.get(invoice_url, headers=headers, params=params) as request:
                        response = await request.content.read()
                        if b"No invoices" in response:
                            self.gui_queue.put(
                                {'status': f'Invoice Not Found : {invoiceno}'}) if self.gui_queue else None
                            self.xldata.append(
                                [self.client, invoiceno[0], invoiceno[1], invoiceno[2], 'Invoice not found'])
                            continue
                        self.xldata.append([self.client, invoiceno[0], invoiceno[1], invoiceno[2],'File downloaded successfully'])
                        file_path = os.path.join(os.getcwd(), "Downloads", self.client)
                        if not os.path.exists(file_path):
                            os.makedirs(file_path)
                        file_name = os.path.join(file_path, f'{invoiceno[0]}.pdf')
                        with open(file_name, 'wb') as f:
                            f.write(response)
                        self.gui_queue.put(
                            {'status': f'Success : {invoiceno[0]}'}) if self.gui_queue else None
        self.gui_queue.put(
            {'status': f'Penn invoices downloded successfully for client {self.client}'}) if self.gui_queue else None
        return True

    async def download_process(self,startdate,enddate):
        timeout = aiohttp.ClientTimeout(total=TIMEOUT)
        conn = aiohttp.TCPConnector(limit=5, limit_per_host=5)
        self.sema = asyncio.Semaphore(LIMIT)
        async with aiohttp.ClientSession(connector=conn, timeout=timeout) as self.session:
            login = await self.login()
            if not login:
                self.gui_queue.put(
                    {'status': f'Username or Password is not valid for client {self.client}'}) \
                    if self.gui_queue else None

            search = await self.search_invoices(startdate, enddate)
            if not search:
                self.gui_queue.put(
                    {'status': f'Invoice Fetch Error: {self.client}'}) \
                    if self.gui_queue else None
            download = await self.download_invoice()
            if not download:
                self.gui_queue.put({'status': f'Download Error: {self.client}'}) \
                    if self.gui_queue else None
            updatexl = await self.update_xl()
            if not updatexl:
                self.gui_queue.put({'status': f'Excel update Error: {self.client}'}) \
                    if self.gui_queue else None


    def start_download(self,startdate,enddate):
        loop = asyncio.new_event_loop()
        executor = ThreadPoolExecutor(max_workers=3)
        future = asyncio.ensure_future(self.download_process(startdate, enddate),loop=loop)
        loop.run_until_complete(future)
        return future.result()

    async def update_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(self.header)
        filename = os.path.join(os.getcwd(), "Downloads", 'Penn.xlsx')
        for xldata in self.xldata:
            ws.append(xldata)
        wb.save(filename)
        return True


class RunPenn:
    def __init__(self):
        self.gui_queue = None

    def run(self,startdate,enddate):
        run_start = time.perf_counter()
        setting = 'PennSettingSheet.xlsx'
        setting_wb = load_workbook(setting, data_only=True, read_only=True)
        setting_ws = setting_wb['Creds'].values
        setting_data = [list(row) for row in setting_ws if row]
        for row_num, row in enumerate(setting_data,1):
            if len(row) >= 3:
                client = str(row[0]).strip()
                username = str(row[1]).strip()
                password = str(row[2]).strip()
                penn = PennInvoice(username, password, client, self.gui_queue)
                penn.start_download(startdate, enddate)

        run_end = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(run_end - run_start)))
        print(f'Time Taken = {time_taken}')
        self.gui_queue.put({"status": f"Time Taken {time_taken}"})

