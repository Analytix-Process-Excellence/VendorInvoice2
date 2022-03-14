import asyncio, aiohttp, os, time
from datetime import date
import json
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup as bs
from concurrent.futures import ThreadPoolExecutor
import queue

LIMIT = 3
TIMEOUT = 600  # seconds
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
             'Chrome/97.0.4692.99 Safari/537.36'

class CovetrusInvoice:
    def __init__(self,gui_queue):
        self.gui_queue = gui_queue
        self.username = self.password = self.client = None
        self.csrf_ = None
        self.invoices = []
        self.doctype = ''
        self.seqno = ''
        self.header = ['Client','DocumentType','InvoiceNumber','OrderDate','SequenceNo','Description']
        self.xldata = []

    async def load_login(self):
        url = 'https://northamerica.covetrus.com/Login?'
        headers = {
            'authority': 'northamerica.covetrus.com',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'dnt': '1',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'none',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-US,en;q=0.9',
        }
        async with self.sema:
            async with self.session.get(url,headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                soup = bs(content,"html.parser")
                self.csrf = soup.find(id="hsv-csrf").get('value')
                self.auth = soup.find(id="hsv-auth-token").get('value')
                # self.cust = soup.find(id="custCount").get('value')
                if self.csrf and self.auth:
                    return True
                else:
                    return False

    async def login(self):
        url = 'https://northamerica.covetrus.com/login/SubmtiCredentials'
        headers = {
            'authority': 'northamerica.covetrus.com',
            'cache-control': 'max-age=0',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'origin': 'https://northamerica.covetrus.com',
            'upgrade-insecure-requests': '1',
            'dnt': '1',
            'content-type': 'application/x-www-form-urlencoded',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-US,en;q=0.9',
        }

        data = {
            'username': self.username,
            'password': self.password,
            'redirectUrl': ''
        }

        async with self.sema:
            async with self.session.post(url, headers=headers, data=data) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                title = html_content.find("title").text
                if str(title).strip() == "Vet Supplies | Pet Medications | Covetrus North America":
                    return True
                else:
                    return False

    async def auth_login(self):
        url = 'https://northamerica.covetrus.com/Login'

        headers = {
            'authority': 'northamerica.covetrus.com',
            'cache-control': 'max-age=0',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        params = (
            ('s', 'SelectCustomer'),
            ('tok', self.auth),
            ('url', '/Default.aspx'),
            ('ExcludeAngular', 'True'),
        )

        async with self.sema:
            async with self.session.get(url, headers=headers, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                print(html_content)
                return True

    async def default_login(self):
        url = 'https://northamerica.covetrus.com/Default.aspx'
        headers = {
            'authority': 'northamerica.covetrus.com',
            'cache-control': 'max-age=0',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        async with self.sema:
            async with self.session.get(url, headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                title = html_content.find("title").text
                if str(title).strip() == "Vet Supplies | Pet Medications | Covetrus North America":
                    self.csrf_ = html_content.find(id="hsv-csrf").get('value')
                    return True
                else:
                    return False

    async def my_account(self):
        url = 'https://northamerica.covetrus.com/My-Account'
        headers = {
            'authority': 'northamerica.covetrus.com',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        async with self.sema:
            async with self.session.get(url, headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                title = html_content.find("title").text
                if str(title).strip() == "My Account":
                    return True
                else:
                    return False

    async def get_invoices(self):
        url = 'https://northamerica.covetrus.com/my-account/orders/invoice-search'

        headers = {
            'authority': 'northamerica.covetrus.com',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="97", "Chromium";v="97"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'referer': 'https://northamerica.covetrus.com/My-Account',
            'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        }

        async with self.sema:
            async with self.session.get(url, headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                title = html_content.find("title").text
                if str(title).strip() == 'Invoice Search':
                    self.invoice_auth = html_content.find(id="hsv-auth-token").get('value')
                    self.invoice_csrf = html_content.find(id="hsv-csrf").get('value')
                    return True
                else:
                    return False

    async def fetch_invoice(self):
        url = 'https://northamerica.covetrus.com/Api/InvoiceApi/GetInvoiceSearch'
        headers = {
            'authority': 'northamerica.covetrus.com',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'dnt': '1',
            'sec-ch-ua-mobile': '?0',
            'user-agent': USER_AGENT,
            'x-hsv-csrf': f'{self.invoice_csrf}',
            'accept': 'application/json, text/javascript, */*; q=0.01',
            'x-requested-with': 'XMLHttpRequest',
            'sec-ch-ua-platform': '"Windows"',
            'x-hsv-authorization': f'{self.invoice_auth}',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'cors',
            'sec-fetch-dest': 'empty',
            'referer': 'https://northamerica.covetrus.com/my-account/orders/invoice-search',
            'accept-language': 'en-US,en;q=0.9',
        }

        params = (
            ('dateFrom', '01/31/2022'),
            ('dateTo', '03/02/2022'),
            ('invoicestatus', '0'),
            ('sortorder', 'InvoiceDate'),
            ('sorttype', 'dsc'),
            ('pagenumber', '1'),
        )

        async with self.sema:
            async with self.session.get(url,headers=headers,params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                self.json_content = json.loads(content)
                for invoice in self.json_content:
                    self.invoices.append([invoice['DocumentType'],invoice['InvoiceNumber'],invoice['SeqNo'],
                                          invoice['OrderDate']])
                return True


    async def download_invoice(self):
        url = 'https://northamerica.covetrus.com/my-account/invoice-detail'

        headers = {
            'authority': 'northamerica.covetrus.com',
            'dnt': '1',
            'upgrade-insecure-requests': '1',
            'user-agent': USER_AGENT,
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Microsoft Edge";v="98"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'none',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'accept-language': 'en-US,en;q=0.9',
        }
        self.file_path = os.path.join(os.getcwd(), "Downloads", self.client, str(date.today()))
        if not os.path.exists(self.file_path):
            os.makedirs(self.file_path)
        for invoices in self.invoices:
            params = (
                    ('RefNo', invoices[1]),
                    ('DocType', invoices[0]),
                    ('seqNo', invoices[2]),
                )
            async with self.sema:
                async with self.session.get(url,headers=headers,params=params) as request:
                    response = await request.content.read()
                    if response:
                        file_name = os.path.join(self.file_path, f'{invoices[1]}.pdf')
                        with open(file_name, 'wb') as f:
                            f.write(response)
                        self.xldata.append([self.client,invoices[0],invoices[1],invoices[3],invoices[2],'File downloaded successfully'])
                        self.gui_queue.put(
                            {'status': f'Success : {invoices[1]}'}) if self.gui_queue else None
                    else:
                        self.xldata.append([self.client,invoices[0],invoices[1],invoices[3],invoices[2],'File not downloaded'])

        self.gui_queue.put(
                    {
                      'status': f'Penn invoices downloded successfully for client {self.client}'}) if self.gui_queue else None
        return True

    async def update_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(self.header)
        filename = os.path.join(os.getcwd(), "Downloads", 'Covetrus.xlsx')
        for xldata in self.xldata:
            ws.append(xldata)
        wb.save(filename)
        return True


    async def download_process(self,startdate,enddate):
        timeout = aiohttp.ClientTimeout(total=TIMEOUT)
        conn = aiohttp.TCPConnector(limit=5, limit_per_host=5)
        self.sema = asyncio.Semaphore(LIMIT)
        async with aiohttp.ClientSession(connector=conn, timeout=timeout) as self.session:
            load = await self.load_login()
            if not load:
                self.gui_queue.put(
                    {'status': f'Webpage load error'}) \
                    if self.gui_queue else None
            login = await self.login()
            if not login:
                self.gui_queue.put(
                    {'status': f'Username or Password is not valid for client {self.client}'}) \
                    if self.gui_queue else None

            search = await self.get_invoices()
            if not search:
                self.gui_queue.put(
                    {'status': f'Error while fetching invoice for client {self.client}'}) \
                    if self.gui_queue else None

            invoice = await self.fetch_invoice()
            if not invoice:
                self.gui_queue.put(
                    {'status': f'Download invoice number error for client {self.client}'}) \
                    if self.gui_queue else None
            download = await self.download_invoice()
            if not download:
                self.gui_queue.put(
                    {'status': f'Download invoice error for client {self.client}'}) \
                    if self.gui_queue else None

            updatexl = await self.update_xl()
            if not updatexl:
                self.gui_queue.put(
                    {'status': f'Excel file error for client {self.client}'}) \
                    if self.gui_queue else None




    def start_download(self,startdate,enddate):
        try:
            loop = asyncio.new_event_loop()
            executor = ThreadPoolExecutor(max_workers=3)
            future = asyncio.ensure_future(self.download_process(startdate,enddate), loop=loop)
            loop.run_until_complete(future)
            return future.result()
        except Exception as e:
            pass



class RunCovetrus:
    def __init__(self):
        self.gui_queue = queue.Queue()

    def run(self,startdate=None,enddate=None):
        run_start = time.perf_counter()
        setting = 'CovetrusSettingSheet.xlsx'
        setting_wb = load_workbook(setting, data_only=True, read_only=True)
        setting_ws = setting_wb['Creds'].values
        setting_data = [list(row) for row in setting_ws if row]
        cov = CovetrusInvoice(self.gui_queue)
        for row_num,row in enumerate(setting_data):
            if len(row) >= 3:
                cov.client = str(row[0]).strip()
                cov.username = str(row[1]).strip()
                cov.password = str(row[2]).strip()
                cov.start_download(startdate, enddate)

        run_end = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(run_end - run_start)))
        print(f'Time Taken = {time_taken}')
        self.gui_queue.put({"status": f"Time Taken {time_taken}"})

if __name__ == '__main__':
    run = RunCovetrus()
    run.run()