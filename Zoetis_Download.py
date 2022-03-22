from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from msedge.selenium_tools import Edge, EdgeOptions
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from openpyxl import load_workbook, Workbook
import os
from time import sleep

class Zoetis:
    def __init__(self):
        self.gui_queue = None
        self.login_url = r'https://www2.zoetisus.com/login'
        self.allfilelist = []
        self.downloadedlist = []


    def start_edge(self, download_pdf=True, download_prompt=False):
        self.downloadPath = os.path.join(os.getcwd(), 'Downloads','Zoetis')
        if not os.path.isdir(self.downloadPath):
            os.makedirs(self.downloadPath)

        self.existing_files = []

        edge_options = EdgeOptions()
        edge_options.use_chromium = True
        edge_options.add_experimental_option(
            "prefs", {
                "behavior": "allow",
                "download.prompt_for_download": download_prompt,
                "plugins.always_open_pdf_externally": download_pdf,
                "download.default_directory": self.downloadPath,
                "safebrowsing.enabled": False,
                "safebrowsing.disable_download_protection": True,
                'profile.default_content_setting_values.automatic_downloads': 1

            }
        )
        self.driver = Edge(
            executable_path=EdgeChromiumDriverManager(log_level=0).install(),
            options=edge_options,
        )
        self.driver.maximize_window()
        return True

    def load_login_page(self):
        self.driver.get(self.login_url)
        trial = 0
        while trial < 3:
            if self.driver.title == "LogIn | Zoetis US":
                sleep(1)
                return True
            else:
                trial += 1
                self.driver.get(self.login_url)
                sleep(2)
        return False

    def popup_check(self):
        popupXpath = '//*[@class="interstitial-redirect__footer-link"]'
        popup = WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.XPATH, popupXpath)))

        cookieXpath = '//*[@id="onetrust-close-btn-container"]//button'
        cookie = WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.XPATH, cookieXpath)))
        sleep(1)
        if cookie:
            cookie.click()
        sleep(1)
        if popup:
            popup.click()
        return True


    def login_zoe(self,username,password,client):

        if not username and not password:
            self.gui_queue.put({'status': f'Credentials not found in setting sheet to download reports.'}) \
                if self.gui_queue else None
            return False

        try:
            usernameXpath = '//*[@name="username" and @id="gigya-loginID-134796880536068290"]'
            user_name = WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.XPATH, usernameXpath)))
            user_name.clear()
            user_name.send_keys(username)

            passwordXpath = '//*[@name="password" and @id="gigya-password-43368134043196960"]'
            password_ = WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.XPATH, passwordXpath)))
            password_.clear()
            password_.send_keys(password)
            sleep(1)

            submitXpath = '//*[contains(@data-screenset-element-id,"__gig_template_element") and @value="Login"]'
            submit = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, submitXpath)))
            submit.click()

            title = "Zoetis Dashboard"
            WebDriverWait(self.driver, 30).until(EC.title_is(title))
            sleep(1)

            accountXpath = '//*[text()="Accounts & Orders"]'
            account = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, accountXpath)))
            account.click()

            sleep(1)
            paymentXpath = '//*[text()="Make a payment"]'
            payment = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, paymentXpath)))
            payment.click()
            sleep(1)


            child_tab = self.driver.window_handles[1]
            self.driver.switch_to.window(child_tab)

            sleep(1)
            acceptXpath = '//*[@id="zz_cnbs_cc_acctrm"]'
            accept = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, acceptXpath)))
            accept.click()
            sleep(1)

            makepaymentXpath = '//*[@class="b_continue" and @value="Accept"]'
            makepayment = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, makepaymentXpath)))
            makepayment.click()
            sleep(1)

            viewinvoiceXpath = '//*[@class="on" and contains(@href,"doc=301")]'
            viewinvoice = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, viewinvoiceXpath)))
            viewinvoice.click()
            sleep(2)


            # sleep(2)
            #
            # downloadXpath = '//*[@alt="Download invoice"]'
            # download = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, downloadXpath)))
            # download.click()
            # sleep(8)
            #
            #
            #

            return True
        except Exception as e:
            print(str(e))
            return False

    def download_invoice(self,startdate,enddate,client):
        fromdateXpath = '//*[@name="m_do_ds_LS_S-DATFR"]'
        fromdate = self.driver.find_element(By.XPATH,fromdateXpath)
        fromdate.clear()
        fromdate.send_keys(startdate)

        todateXpath = '//*[@name="m_do_ds_LS_S-DATTO"]'
        todate = self.driver.find_element(By.XPATH,todateXpath)
        todate.clear()
        todate.send_keys(enddate)

        searchXpath = '//*[@class="b_search"]'
        search = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, searchXpath)))
        search.click()


        itemsXpath = '//*[@name="m_do_dl_GV_MAXVIEW" and @class="prinputDef7"]'
        items = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, itemsXpath)))
        items.click()

        itemsXpath = '//*[@name="m_do_dl_GV_MAXVIEW" and @class="prinputDef7"]/option[@value="9999"]'
        items = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, itemsXpath)))
        items.click()

        sleep(2)

        docnum1Xpath = '//*[@class="contentrow2"]//td[3]'
        docnum2Xpath = '//*[@class="contentrow1"]//td[3]'
        docnums1 = self.driver.find_elements(By.XPATH,docnum1Xpath)
        docnums2 = self.driver.find_elements(By.XPATH, docnum2Xpath)

        for docnum1 in docnums1:
            self.allfilelist.append([client,f"Zoetis-invoices-{docnum1.get_attribute('innerHTML')}"])

        for docnum2 in docnums2:
            self.allfilelist.append([client,f"Zoetis-invoices-{docnum2.get_attribute('innerHTML')}"])

        invoicesXpath = '//*[@alt="Download invoice"]'
        invoices = self.driver.find_elements(By.XPATH,invoicesXpath)

        for invoice in invoices:
            try:
                self.driver.execute_script("arguments[0].click();", invoice)
                sleep(0.5)
            except Exception as e:
                print(str(e))

        parent_tab = self.driver.window_handles[0]
        self.driver.close()
        self.driver.switch_to.window(parent_tab)
        return True


    def logout(self):
        accountXpath = '//*[text()="Accounts & Orders"]'
        account = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, accountXpath)))
        account.click()

        logoutXpath = '/html/body/header/div/div/div/div/div[3]/div[3]/ul/li/div/div[2]/div[2]/a/span'
        logout = WebDriverWait(self.driver, 30).until(EC.element_to_be_clickable((By.XPATH, logoutXpath)))
        logout.click()

        logout_title = "Zoetis | zoetisUS.com | Zoetis US"
        logout_flag = WebDriverWait(self.driver, 30).until(EC.title_is((logout_title)))
        if logout_flag:
            sleep(1)
            return True
        else:
            return False



    def get_download_list(self):
        downloadedfiles = os.listdir(self.downloadPath)
        files = [x.split('.')[0] for x in downloadedfiles]
        summary_wb = Workbook()
        summary_ws = summary_wb.active

        for data in self.allfilelist:
            if data[1] not in files:
                summary_ws.append([data[0],data[1],"Not Downloaded"])
            else:
                summary_ws.append([data[0],data[1],"Downloaded Sucecssfully"])
        filepath = os.path.join(self.downloadPath,'ZoetisSummary.xlsx')
        summary_wb.save(filepath)
        return True


class RunZoetis:
    def __init__(self):
        self.gui_queue = None

    def run(self):
        startdate = '02/01/2022' #MMDDYYYY
        enddate = '02/28/2022'
        setting = 'ZoetisSettingSheet.xlsx'
        setting_wb = load_workbook(setting, data_only=True, read_only=True)
        setting_ws = setting_wb['Creds'].values
        setting_data = [list(row) for row in setting_ws if row]
        zoetis = Zoetis()
        edge = zoetis.start_edge()
        if not edge:
            self.gui_queue.put({'status': f'\nError : Unable to load browser.'}) if self.gui_queue else None
            return False
        for row_num, row in enumerate(setting_data,1):
            if len(row) >= 3:
                client = str(row[0]).strip()
                username = str(row[1]).strip()
                password = str(row[2]).strip()
                sleep(2)
                login_page = zoetis.load_login_page()
                if not login_page:
                    self.gui_queue.put({'status': f'\nError : Unable to load login page.'}) if self.gui_queue else None
                    return False
                if row_num == 1:
                    popup = zoetis.popup_check()
                    if not popup:
                        self.gui_queue.put({'status': f'\nError : Unable to close popup.'}) if self.gui_queue else None
                        return False
                login = zoetis.login_zoe(username,password,client)
                if not login:
                    self.gui_queue.put({'status': f'\nError : Unable to login.'}) if self.gui_queue else None
                    return False
                download = zoetis.download_invoice(startdate,enddate,client)
                if not download:
                    self.gui_queue.put({'status': f'\nError : Unable to download invoice for client {client}.'}) if self.gui_queue else None
                    return False
                logout = zoetis.logout()
                if not logout:
                    self.gui_queue.put({'status': f'\nError : Unable to logout.'}) if self.gui_queue else None
                    return False
            sleep(2)
        zoetis.driver.quit()
        filelist = zoetis.get_download_list()
        if not filelist:
            self.gui_queue.put(
                {'status': f'\nError : Unable to get file not downloaded list.'}) if self.gui_queue else None
            return False




if __name__ == '__main__':
    zoe = RunZoetis()
    zoe.run()

